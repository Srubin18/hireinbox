#!/usr/bin/env python3
"""
HIREINBOX 18-Month Financial Forecast Generator
Creates a professional Excel spreadsheet for investor/partner presentations
All pricing is ex VAT
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# Output path
OUTPUT_PATH = "/Users/simon/Desktop/hireinbox/HIREINBOX_18_Month_Forecast.xlsx"

# Create workbook
wb = Workbook()

# Styles
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
subheader_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
money_positive = Font(color="006400")
money_negative = Font(color="8B0000")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
currency_format = 'R #,##0'
currency_format_neg = 'R #,##0;[Red]-R #,##0'

def style_header_row(ws, row_num, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def auto_column_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width

# ============== SHEET 1: SUMMARY ==============
ws_summary = wb.active
ws_summary.title = "Summary"

# Title
ws_summary['A1'] = "HIREINBOX - 18 Month Financial Forecast"
ws_summary['A1'].font = Font(bold=True, size=16)
ws_summary['A2'] = "February 2026 - July 2027 | All figures ex VAT"
ws_summary['A2'].font = Font(italic=True, size=11)
ws_summary['A3'] = f"Generated: {datetime.now().strftime('%d %B %Y')}"

# Key Metrics
ws_summary['A5'] = "KEY METRICS"
ws_summary['A5'].font = Font(bold=True, size=14)

metrics = [
    ["Metric", "Value"],
    ["Total Investment Required", "R 5,500,000"],
    ["Build Phase Cost", "R 1,358,000"],
    ["Monthly Burn (Post-Launch)", "R 404,000"],
    ["Break-Even Month", "Month 17 (Jun 2027)"],
    ["Month 18 Revenue", "R 485,000"],
    ["Month 18 Profit", "R 81,000"],
]

for i, row in enumerate(metrics):
    ws_summary.cell(row=6+i, column=1, value=row[0]).border = thin_border
    ws_summary.cell(row=6+i, column=2, value=row[1]).border = thin_border
    if i == 0:
        ws_summary.cell(row=6+i, column=1).font = Font(bold=True)
        ws_summary.cell(row=6+i, column=2).font = Font(bold=True)

# ============== SHEET 2: MONTHLY FORECAST ==============
ws_monthly = wb.create_sheet("Monthly Forecast")

# Headers
headers = ["Month", "Date", "Phase", "Expenses", "B2B Revenue", "B2C Revenue", "Total Revenue", "Net P/L", "Cumulative"]
for col, header in enumerate(headers, 1):
    ws_monthly.cell(row=1, column=col, value=header)
style_header_row(ws_monthly, 1, len(headers))

# Data
monthly_data = [
    [1, "Feb 2026", "Build", 433333, 0, 0, 0, -433333, -433333],
    [2, "Mar 2026", "Build", 378333, 0, 0, 0, -378333, -811666],
    [3, "Apr 2026", "Build", 546333, 0, 0, 0, -546333, -1358000],
    [4, "May 2026", "Launch", 324000, 8750, 3000, 11750, -312250, -1670250],
    [5, "Jun 2026", "Grow", 324000, 17500, 6000, 23500, -300500, -1970750],
    [6, "Jul 2026", "Grow", 404000, 43750, 14415, 58165, -345835, -2316585],
    [7, "Aug 2026", "Grow", 404000, 61250, 18000, 79250, -324750, -2641335],
    [8, "Sep 2026", "Grow", 404000, 78750, 22000, 100750, -303250, -2944585],
    [9, "Oct 2026", "Grow", 404000, 105000, 28000, 133000, -271000, -3215585],
    [10, "Nov 2026", "Grow", 404000, 131250, 35000, 166250, -237750, -3453335],
    [11, "Dec 2026", "Grow", 404000, 148750, 40000, 188750, -215250, -3668585],
    [12, "Jan 2027", "Grow", 404000, 175000, 48000, 223000, -181000, -3849585],
    [13, "Feb 2027", "Grow", 404000, 201250, 55000, 256250, -147750, -3997335],
    [14, "Mar 2027", "Grow", 404000, 227500, 62000, 289500, -114500, -4111835],
    [15, "Apr 2027", "Grow", 404000, 262500, 70000, 332500, -71500, -4183335],
    [16, "May 2027", "Grow", 404000, 297500, 78000, 375500, -28500, -4211835],
    [17, "Jun 2027", "Grow", 404000, 341250, 88000, 429250, 25250, -4186585],
    [18, "Jul 2027", "Grow", 404000, 385000, 100000, 485000, 81000, -4105585],
]

for row_idx, row_data in enumerate(monthly_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_monthly.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        # Format currency columns
        if col_idx >= 4:
            cell.number_format = currency_format_neg
        # Highlight break-even row
        if row_data[0] == 17:
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

auto_column_width(ws_monthly)

# ============== SHEET 3: B2B PRICING ==============
ws_b2b = wb.create_sheet("B2B Pricing (ex VAT)")

ws_b2b['A1'] = "B2B PRICING - All prices ex VAT"
ws_b2b['A1'].font = Font(bold=True, size=14)

b2b_headers = ["Product", "Price", "Unit", "Description"]
for col, header in enumerate(b2b_headers, 1):
    ws_b2b.cell(row=3, column=col, value=header)
style_header_row(ws_b2b, 3, len(b2b_headers))

b2b_products = [
    ["CV Screening", "R 1,750", "per role", "Unlimited CVs per role, AI scoring & ranking"],
    ["ID Verification", "R 50", "per candidate", "Verify candidate identity"],
    ["Credit Check", "R 100", "per candidate", "Financial background verification"],
    ["Criminal Check", "R 150", "per candidate", "Criminal record verification"],
    ["AI Interview + Psychometric", "R 750", "per candidate", "Avatar interview with full analysis"],
    ["Job Listing (Phase 2)", "R 2,500", "per listing", "Post job publicly to attract candidates"],
    ["Subscription Starter (Phase 3)", "R 5,000", "per month", "Up to 10 roles/month"],
    ["Subscription Pro (Phase 3)", "R 10,000", "per month", "Up to 25 roles/month"],
    ["Subscription Enterprise (Phase 3)", "R 15,000", "per month", "Unlimited roles + support"],
    ["Boutique AI Agent", "R 20,000", "per month", "Custom-trained AI for your company"],
]

for row_idx, row_data in enumerate(b2b_products, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_b2b.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

auto_column_width(ws_b2b)

# ============== SHEET 4: B2C PRICING ==============
ws_b2c = wb.create_sheet("B2C Pricing (ex VAT)")

ws_b2c['A1'] = "B2C PRICING - All prices ex VAT"
ws_b2c['A1'].font = Font(bold=True, size=14)

b2c_headers = ["Product", "Price", "Description"]
for col, header in enumerate(b2c_headers, 1):
    ws_b2c.cell(row=3, column=col, value=header)
style_header_row(ws_b2c, 3, len(b2c_headers))

b2c_products = [
    ["CV Scan", "FREE (1x)", "AI analysis of CV with feedback"],
    ["CV Redo/Rewrite", "FREE (1x)", "AI rewrites CV professionally"],
    ["Video Analysis", "R 149", "AI coaching on interview video"],
    ["AI Avatar Coaching", "R 199", "Interview prep with AI avatar"],
    ["Position-Specific Prep", "R 199", "Guidance for specific job application"],
    ["Video Pitch Package", "R 149", "Create video pitch for employers"],
]

for row_idx, row_data in enumerate(b2c_products, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_b2c.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

auto_column_width(ws_b2c)

# ============== SHEET 5: TEAM & SALARIES ==============
ws_team = wb.create_sheet("Team & Salaries")

ws_team['A1'] = "TEAM & SALARIES - Cape Town Startup Market Rates"
ws_team['A1'].font = Font(bold=True, size=14)

team_headers = ["Role", "Monthly Salary", "Start Date", "Notes"]
for col, header in enumerate(team_headers, 1):
    ws_team.cell(row=3, column=col, value=header)
style_header_row(ws_team, 3, len(team_headers))

team_data = [
    ["Marketing Manager", "R 45,000", "1 Apr 2026", "Mid-senior, growth-focused"],
    ["Full-Stack Developer", "R 60,000", "1 Apr 2026", "Senior, Cape Town rate"],
    ["Success Manager", "R 38,000", "1 Apr 2026", "Mid-level, customer-focused"],
    ["CEO (Simon Rubin)", "R 40,000", "1 Jul 2026", "Below market, founder"],
    ["Co-CEO (Shay Sinbeti)", "R 40,000", "1 Jul 2026", "Below market, founder"],
]

for row_idx, row_data in enumerate(team_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_team.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

ws_team['A10'] = "Team hired 1 month before launch to allow for onboarding and preparation."
ws_team['A10'].font = Font(italic=True)

auto_column_width(ws_team)

# ============== SHEET 6: EXPENSES BREAKDOWN ==============
ws_expenses = wb.create_sheet("Expense Breakdown")

ws_expenses['A1'] = "MONTHLY EXPENSE BREAKDOWN (Post-Launch)"
ws_expenses['A1'].font = Font(bold=True, size=14)

expense_headers = ["Category", "Item", "Monthly Cost"]
for col, header in enumerate(expense_headers, 1):
    ws_expenses.cell(row=3, column=col, value=header)
style_header_row(ws_expenses, 3, len(expense_headers))

expenses = [
    ["Salaries", "Marketing Manager", 45000],
    ["Salaries", "Full-Stack Developer", 60000],
    ["Salaries", "Success Manager", 38000],
    ["Salaries", "CEO (from Month 6)", 40000],
    ["Salaries", "Co-CEO (from Month 6)", 40000],
    ["Marketing", "Advertising & Campaigns", 100000],
    ["Technology", "Cloud Hosting", 10000],
    ["Technology", "AI API (OpenAI)", 20000],
    ["Technology", "Database (Supabase)", 3000],
    ["Technology", "Tools & Services", 5000],
    ["Operations", "Office/Co-working", 15000],
    ["Operations", "Insurance", 5000],
    ["Operations", "Accounting", 8000],
    ["Operations", "Legal", 5000],
    ["Operations", "Miscellaneous", 10000],
]

for row_idx, row_data in enumerate(expenses, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_expenses.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if col_idx == 3:
            cell.number_format = currency_format

# Total row
total_row = len(expenses) + 4
ws_expenses.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
ws_expenses.cell(row=total_row, column=3, value=404000).font = Font(bold=True)
ws_expenses.cell(row=total_row, column=3).number_format = currency_format

auto_column_width(ws_expenses)

# Save workbook
wb.save(OUTPUT_PATH)
print(f"Excel file created: {OUTPUT_PATH}")
print("Sheets: Summary, Monthly Forecast, B2B Pricing, B2C Pricing, Team & Salaries, Expense Breakdown")
